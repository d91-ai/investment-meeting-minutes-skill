#!/usr/bin/env python3
"""Query local stock-symbol resources and return ranked candidates."""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from dataclasses import asdict, dataclass
from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path
from typing import Iterable

try:
    from pypinyin import Style, lazy_pinyin
except ImportError:  # Optional until phonetic recall is explicitly required.
    Style = None  # type: ignore[assignment]
    lazy_pinyin = None  # type: ignore[assignment]

DEFAULT_WORKSPACE_ROOT = (
    Path(os.environ["INVESTMENT_MINUTES_WORKSPACE"]).expanduser()
    if os.environ.get("INVESTMENT_MINUTES_WORKSPACE")
    else Path.home() / "Documents/会议纪要整理"
)
DEFAULT_SYMBOL_ROOT = DEFAULT_WORKSPACE_ROOT / "03 Resources/market-symbols"
DEFAULT_ALIAS_PATH = DEFAULT_SYMBOL_ROOT / "company_aliases.csv"
CJK_PATTERN = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")


@dataclass
class Candidate:
    symbol: str
    name: str
    market: str
    confidence: float
    match_type: str
    source: str
    phonetic_similarity: float | None = None


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", "", value).strip().upper()


def normalize_code(value: str) -> str:
    return value.strip().upper().replace(" ", "")


def phonetic_available() -> bool:
    return lazy_pinyin is not None and Style is not None


def normalize_phonetic_source(value: str) -> str:
    cleaned = re.sub(r"[（(][0-9A-Z.]+[)）]\s*$", "", value.strip(), flags=re.IGNORECASE)
    return re.sub(r"[-‐‑–—](?:U|W|B|SW|UW)\s*$", "", cleaned, flags=re.IGNORECASE)


@lru_cache(maxsize=32768)
def phonetic_forms(value: str) -> tuple[str, str] | None:
    source = normalize_phonetic_source(value)
    if not source or not CJK_PATTERN.search(source) or not phonetic_available():
        return None
    syllables = [
        re.sub(r"[^a-z0-9]", "", item.lower().replace("ü", "v"))
        for item in lazy_pinyin(source, style=Style.NORMAL, errors=lambda chars: list(chars))  # type: ignore[misc,union-attr]
    ]
    syllables = [item for item in syllables if item]
    if not syllables:
        return None
    return "".join(syllables), "".join(item[0] for item in syllables)


def phonetic_match_forms(
    query_forms: tuple[str, str],
    target_forms: tuple[str, str],
) -> tuple[float, str, float] | None:
    query_full, query_initials = query_forms
    target_full, target_initials = target_forms
    if query_full == target_full:
        return 0.92, "pinyin_exact", 1.0
    similarity = SequenceMatcher(None, query_full, target_full).ratio()
    if min(len(query_full), len(target_full)) >= 4 and similarity >= 0.82:
        return round(0.55 + similarity * 0.3, 3), "pinyin_fuzzy", round(similarity, 3)
    if len(query_initials) >= 2 and query_initials == target_initials:
        return 0.62, "pinyin_initials", 1.0
    return None


def phonetic_match(query: str, target: str) -> tuple[float, str, float] | None:
    query_forms = phonetic_forms(query)
    target_forms = phonetic_forms(target)
    if query_forms is None or target_forms is None:
        return None
    return phonetic_match_forms(query_forms, target_forms)


def load_a_share(root: Path) -> list[dict[str, str]]:
    path = root / "a_share_list.csv"
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [
            {
                "symbol": row.get("symbol", "").strip(),
                "code": row.get("code", "").strip(),
                "name": row.get("name", "").strip(),
                "market": "A",
                "source": str(path),
            }
            for row in csv.DictReader(handle)
            if row.get("symbol") and row.get("name")
        ]


def load_us(root: Path) -> list[dict[str, str]]:
    source_dir = root / "global-stock-symbols-source"
    files = [
        ("NASDAQ", source_dir / "nasadaq_3105.csv"),
        ("NYSE", source_dir / "nyse_3105.csv"),
        ("AMEX", source_dir / "amex_3105.csv"),
    ]
    rows: list[dict[str, str]] = []
    for market, path in files:
        if not path.exists():
            continue
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                symbol = (row.get("Symbol") or "").strip()
                name = (row.get("Name") or "").strip()
                if symbol and name:
                    rows.append({"symbol": symbol, "code": symbol, "name": name, "market": market, "source": str(path)})
    return rows


def load_hk(root: Path) -> list[dict[str, str]]:
    path = root / "global-stock-symbols-source/HongLongListOfSecurities_300625.xlsx"
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    rows: list[dict[str, str]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header_seen = False
    for values in sheet.iter_rows(values_only=True):
        if not header_seen:
            header_seen = values and values[0] == "Stock Code"
            continue
        code = str(values[0] or "").strip()
        name = str(values[1] or "").strip()
        category = str(values[2] or "").strip()
        if not code or not name or category != "Equity":
            continue
        rows.append({"symbol": f"{code}.HK", "code": code, "name": name, "market": "HK", "source": str(path)})
    return rows


def load_aliases(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle) if row.get("query") and row.get("symbol")]


def load_session_entities(path: Path | None) -> list[dict[str, object]]:
    if path is None:
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    items = payload.get("entities") if isinstance(payload, dict) else None
    if not isinstance(items, list):
        raise ValueError("session entities 必须是包含 entities 数组的 JSON object")
    entities: list[dict[str, object]] = []
    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"entities[{index}] 必须是 JSON object")
        name = str(item.get("name") or "").strip()
        symbol = normalize_code(str(item.get("symbol") or ""))
        aliases = item.get("aliases", [])
        if not name or not symbol or not isinstance(aliases, list) or any(not isinstance(alias, str) for alias in aliases):
            raise ValueError(f"entities[{index}] 必须提供 name、symbol 和字符串 aliases 数组")
        entities.append(
            {
                "name": name,
                "symbol": symbol,
                "market": str(item.get("market") or "").strip(),
                "aliases": [alias.strip() for alias in aliases if alias.strip()],
            }
        )
    return entities


def load_symbol_rows(root: Path) -> list[dict[str, str]]:
    return load_a_share(root) + load_hk(root) + load_us(root)


def market_from_symbol(symbol: str, declared: str) -> str:
    if declared:
        return declared
    if symbol.endswith((".SH", ".SZ", ".BJ")):
        return "A"
    if symbol.endswith(".HK"):
        return "HK"
    return "US" if re.fullmatch(r"[A-Z.]+", symbol) else "session"


def candidate_from_session(
    query: str,
    entities: list[dict[str, object]],
    *,
    source: str,
) -> list[Candidate]:
    normalized_query = normalize_text(query)
    candidates: list[Candidate] = []
    for entity in entities:
        name = str(entity["name"])
        symbol = normalize_code(str(entity["symbol"]))
        market = market_from_symbol(symbol, str(entity.get("market") or ""))
        forms = [name, *[str(alias) for alias in entity.get("aliases", [])]]
        for form in forms:
            normalized_form = normalize_text(form)
            similarity: float | None = None
            if normalized_query == normalized_form:
                confidence, match_type = 0.995, "session_entity_exact"
            elif normalized_query and (normalized_query in normalized_form or normalized_form in normalized_query):
                confidence, match_type = 0.9, "session_entity_alias"
            else:
                phonetic = phonetic_match(query, form)
                if phonetic is None:
                    continue
                confidence, phonetic_type, similarity = phonetic
                confidence = min(0.96, round(confidence + 0.03, 3))
                match_type = f"session_entity_{phonetic_type}"
            candidates.append(Candidate(symbol, name, market, confidence, match_type, source, similarity))
    return candidates


def candidate_from_alias(query: str, aliases: list[dict[str, str]], *, alias_path: Path) -> list[Candidate]:
    normalized_query = normalize_text(query)
    candidates: list[Candidate] = []
    for row in aliases:
        alias = normalize_text(row.get("query", ""))
        name = row.get("name", "").strip() or row.get("query", "").strip()
        if normalized_query == alias:
            confidence = float(row.get("confidence") or 0.98)
            match_type = "alias_exact"
        elif normalized_query and (normalized_query in alias or alias in normalized_query):
            confidence = float(row.get("partial_confidence") or 0.86)
            match_type = "alias_partial"
        else:
            phonetic = phonetic_match(query, row.get("query", ""))
            if phonetic is None:
                continue
            confidence, match_type, similarity = phonetic
        candidates.append(
            Candidate(
                symbol=normalize_code(row.get("symbol", "")),
                name=name,
                market=(row.get("market") or "alias").strip(),
                confidence=confidence,
                match_type=match_type,
                source=str(alias_path),
                phonetic_similarity=similarity if match_type.startswith("pinyin_") else None,
            )
        )
    return candidates


def score_row(query: str, row: dict[str, str]) -> Candidate | None:
    raw_query = query.strip()
    normalized_query = normalize_text(raw_query)
    symbol = normalize_code(row["symbol"])
    code = normalize_code(row.get("code", ""))
    name = row["name"].strip()
    normalized_name = normalize_text(name)

    if not normalized_query:
        return None
    if normalize_code(raw_query) in {symbol, code}:
        return Candidate(symbol, name, row["market"], 1.0, "code_exact", row["source"])
    if normalized_query == normalized_name:
        return Candidate(symbol, name, row["market"], 0.98, "name_exact", row["source"])
    if normalized_query in normalized_name or normalized_name in normalized_query:
        return Candidate(symbol, name, row["market"], 0.88, "name_contains", row["source"])
    ratio = SequenceMatcher(None, normalized_query, normalized_name).ratio()
    if ratio >= 0.72:
        return Candidate(symbol, name, row["market"], round(0.55 + ratio * 0.3, 3), "name_fuzzy", row["source"])
    phonetic = phonetic_match(raw_query, name)
    if phonetic is not None:
        confidence, match_type, similarity = phonetic
        return Candidate(symbol, name, row["market"], confidence, match_type, row["source"], similarity)
    return None


def market_allowed(row_market: str, requested: str) -> bool:
    if requested == "all":
        return True
    if requested == "US":
        return row_market in {"NASDAQ", "NYSE", "AMEX", "US"}
    return row_market == requested


def dedupe(candidates: Iterable[Candidate]) -> list[Candidate]:
    best: dict[str, Candidate] = {}
    for candidate in candidates:
        current = best.get(candidate.symbol)
        if current is None or candidate.confidence > current.confidence:
            best[candidate.symbol] = candidate
    return sorted(best.values(), key=lambda item: (-item.confidence, item.market, item.symbol))


def query_symbols(
    query: str,
    *,
    market: str,
    limit: int,
    root: Path,
    alias_path: Path,
    rows: list[dict[str, str]] | None = None,
    aliases: list[dict[str, str]] | None = None,
    session_entities: list[dict[str, object]] | None = None,
    session_source: str = "current_session",
) -> dict[str, object]:
    rows = rows if rows is not None else load_symbol_rows(root)
    aliases = aliases if aliases is not None else load_aliases(alias_path)
    candidates: list[Candidate] = []
    candidates.extend(candidate_from_session(query, session_entities or [], source=session_source))
    candidates.extend(candidate_from_alias(query, aliases, alias_path=alias_path))
    for row in rows:
        if not market_allowed(row["market"], market):
            continue
        candidate = score_row(query, row)
        if candidate:
            candidates.append(candidate)

    ranked = dedupe(candidate for candidate in candidates if market_allowed(candidate.market, market))[:limit]
    status = "candidate_only" if ranked else "not_found"
    return {
        "query": query,
        "market": market,
        "status": status,
        "confirmed": False,
        "recommendation": None,
        "phonetic_backend": "pypinyin" if phonetic_available() else "unavailable",
        "candidates": [asdict(candidate) for candidate in ranked],
    }


def read_batch_queries(path: Path) -> list[str]:
    queries: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        item = line.strip()
        if item and not item.startswith("#"):
            queries.append(item)
    return queries


def print_batch_text(payloads: list[dict[str, object]]) -> None:
    for index, payload in enumerate(payloads):
        if index:
            print()
        print_text(payload)


def print_text(payload: dict[str, object]) -> None:
    print(f"查询: {payload['query']}")
    print(f"状态: {payload['status']}")
    for item in payload["candidates"]:  # type: ignore[index]
        print(
            f"- {item['symbol']} | {item['name']} | {item['market']} | "
            f"{item['confidence']:.2f} | {item['match_type']}"
        )
    if payload["status"] == "candidate_only":
        print("处理建议: 仅作候选线索；使用外部证据核验后才能写入确认代码，未核验前保留存疑。")
    else:
        print("处理建议: 不要直接写入代码；放入“存疑与待确认”或补充人工校对参考。")


def main() -> int:
    parser = argparse.ArgumentParser(description="本地查询股票代码候选和置信度")
    parser.add_argument("query", nargs="?", help="公司名、简称、股票代码或 ticker")
    parser.add_argument("--query", dest="extra_queries", action="append", default=[], help="追加查询词，可重复")
    parser.add_argument("--batch-file", help="批量查询文件，每行一个公司名、简称、股票代码或 ticker")
    parser.add_argument("--market", choices=["all", "A", "HK", "US", "NASDAQ", "NYSE", "AMEX"], default="all")
    parser.add_argument("--limit", type=int, default=8)
    parser.add_argument("--root", default=str(DEFAULT_SYMBOL_ROOT), help="market-symbols 目录")
    parser.add_argument("--aliases", default=str(DEFAULT_ALIAS_PATH), help="公司别名 CSV")
    parser.add_argument("--session-entities", help="当前会议已确认实体 JSON，仅用于本次候选召回")
    parser.add_argument("--require-phonetic", action="store_true", help="拼音召回不可用时阻断查询")
    parser.add_argument("--json", action="store_true", help="输出 JSON")
    args = parser.parse_args()

    if args.require_phonetic and not phonetic_available():
        print("拼音召回不可用；请安装 skills/投资会议纪要整理/requirements.txt", file=sys.stderr)
        return 2

    queries: list[str] = []
    if args.query:
        queries.append(args.query)
    queries.extend(args.extra_queries)
    if args.batch_file:
        queries.extend(read_batch_queries(Path(args.batch_file).expanduser()))
    queries = [item.strip() for item in queries if item.strip()]
    if not queries:
        parser.error("请提供 query、--query 或 --batch-file")

    root = Path(args.root).expanduser()
    alias_path = Path(args.aliases).expanduser()
    session_path = Path(args.session_entities).expanduser() if args.session_entities else None
    rows = load_symbol_rows(root)
    aliases = load_aliases(alias_path)
    try:
        session_entities = load_session_entities(session_path)
    except (OSError, UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
        print(f"无法读取当前会议实体: {exc}", file=sys.stderr)
        return 2
    payloads = [
        query_symbols(
            query,
            market=args.market,
            limit=args.limit,
            root=root,
            alias_path=alias_path,
            rows=rows,
            aliases=aliases,
            session_entities=session_entities,
            session_source=str(session_path) if session_path else "current_session",
        )
        for query in queries
    ]
    if args.json:
        if len(payloads) == 1:
            print(json.dumps(payloads[0], ensure_ascii=False, indent=2))
        else:
            print(
                json.dumps(
                    {
                        "query_count": len(payloads),
                        "ok": all(payload["status"] != "not_found" for payload in payloads),
                        "results": payloads,
                    },
                    ensure_ascii=False,
                    indent=2,
                )
            )
    else:
        print_batch_text(payloads)
    return 0 if all(payload["candidates"] for payload in payloads) else 1


if __name__ == "__main__":
    raise SystemExit(main())
